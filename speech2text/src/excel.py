from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

from speech2text.src.models import RuleID, Report


class Excel:
    FILE = Path('Звіт прослуханих розмов.xlsx')

    # Создал эту переменную только потому что не получается нормально
    #   найти последнюю пустую строку в Excel файле
    #   из-за форматирования и пропущенных строк
    START_ROW = 1341

    def __init__(self):
        self.wb: Workbook | None = None
        self.ws: "Worksheet | None" = None

        # Временное решение для нахождения пустой строки
        self.i = 0

    def __enter__(self):
        self.wb = load_workbook(self.FILE)
        self.ws = self.wb.active
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.wb.save(Excel.FILE)
        self.wb.close()

    @property
    def next_row(self):
        # for index, row in enumerate(self.ws.rows, 1):
        #     if all(cell.value is None for cell in row):
        #         return index
        # return self.ws.max_row + 1

        # Временное решение для нахождения пустой строки
        self.i += 1
        return Excel.START_ROW + self.i

    def write_report(self, report: Report):
        """
        Записываем наш фидбек в таблицу
        """
        row = self.next_row

        # Дата
        self.ws[f'A{row}'] = report.date

        # Запись
        self.ws[f'L{row}'] = report.appointment

        check_list = {
            item.id: item.check
            for item in report.result
        }

        # Чек-лист
        self.ws[f'F{row}'] = int(check_list[RuleID.GREETING])
        self.ws[f'G{row}'] = int(check_list[RuleID.CAR_BODY])
        self.ws[f'H{row}'] = int(check_list[RuleID.CAR_YEAR])
        self.ws[f'I{row}'] = int(check_list[RuleID.CAR_MILEAGE])
        self.ws[f'J{row}'] = int(check_list[RuleID.COMPREHENSIVE_DIAGNOSIS])
        self.ws[f'K{row}'] = int(check_list[RuleID.PREVIOUS_WORKS])
        self.ws[f'M{row}'] = int(check_list[RuleID.PARTING])

        # Оценка
        self.ws[f'R{row}'] = sum([i for i in check_list.values()])

        # Комментарий
        self.ws[f'T{row}'] = report.bad_moments + ' ' + report.overall

        if report.bad_moments:
            self.ws[f'T{row}'].fill = PatternFill(
                fill_type='solid',
                fgColor='FF0000'
            )
